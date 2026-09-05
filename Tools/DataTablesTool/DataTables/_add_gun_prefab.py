# -*- coding: utf-8 -*-
"""Add gun_prefab column to #Weapon.xlsx and rebuild Datas_Gen."""
from __future__ import annotations

import os
import shutil
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent
DATAS = ROOT / "Datas"
DATAS_GEN = ROOT / "Datas_Gen"

GUN_BY_ID = {
    0: "Gun-Rifle",
    1: "Gun-Shotgun",
    2: "Gun-RPG",
    3: "Gun-Energy",
    4: "Gun-Sniper",
    5: "Gun-Crossbow",
    6: "EnemyGun_Canon_A",
}


def ensure_gun_prefab(path: Path) -> None:
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    if "gun_prefab" in headers:
        col = headers.index("gun_prefab") + 1
    else:
        # insert after weapon_key (col D index 4) -> after name block: find weapon_key
        # append before in_shop if present, else at end
        if "in_shop" in headers:
            insert_at = headers.index("in_shop") + 1  # 1-based excel col for insert
        else:
            insert_at = len(headers) + 1
        ws.insert_cols(insert_at)
        ws.cell(1, insert_at).value = "gun_prefab"
        ws.cell(2, insert_at).value = "string"
        ws.cell(3, insert_at).value = "c"
        ws.cell(4, insert_at).value = "枪械逻辑Prefab location"
        col = insert_at
        # fix ##var first cell
        ws.cell(1, 1).value = "##var"
        ws.cell(2, 1).value = "##type"
        ws.cell(3, 1).value = "##group"
        ws.cell(4, 1).value = "##"

    # find id column
    headers = [c.value for c in ws[1]]
    id_col = headers.index("id") + 1
    gun_col = headers.index("gun_prefab") + 1 if "gun_prefab" in headers else col

    for r in range(5, ws.max_row + 1):
        wid = ws.cell(r, id_col).value
        if wid is None:
            continue
        ws.cell(r, gun_col).value = GUN_BY_ID.get(int(wid), "")

    tmp = path.with_suffix(".xlsx.tmp")
    wb.save(tmp)
    try:
        os.replace(tmp, path)
        print(f"updated {path}")
    except PermissionError:
        alt = path.with_name(path.stem + ".with_gun.xlsx")
        os.replace(tmp, alt)
        print(f"locked, wrote {alt}")


def rebuild_datas_gen() -> None:
    if DATAS_GEN.exists():
        shutil.rmtree(DATAS_GEN)
    DATAS_GEN.mkdir()
    for f in DATAS.iterdir():
        if not f.is_file() or f.name.startswith("~$"):
            continue
        if f.name.startswith("#Weapon") and f.suffix == ".xlsx":
            continue
        shutil.copy2(f, DATAS_GEN / f.name)

    # prefer updated Datas/#Weapon.xlsx, else with_gun, else Datas_Gen source from previous
    candidates = [
        DATAS / "#Weapon.xlsx",
        DATAS / "#Weapon.with_gun.xlsx",
        ROOT / "Datas_Gen_backup_skip",
    ]
    src = None
    for c in candidates:
        if c.exists():
            src = c
            break
    if src is None:
        # update in-memory from filled schema in script by copying Datas_Gen old if any
        raise SystemExit("no weapon xlsx source")

    # If Datas #Weapon was locked we may have written with_gun; always ensure gun column
    ensure_gun_prefab(src)
    shutil.copy2(src, DATAS_GEN / "#Weapon.xlsx")
    ensure_gun_prefab(DATAS_GEN / "#Weapon.xlsx")
    print("Datas_Gen ready")


def main() -> None:
    # try update real Datas
    weapon = DATAS / "#Weapon.xlsx"
    if weapon.exists():
        try:
            ensure_gun_prefab(weapon)
        except SystemExit:
            pass
    rebuild_datas_gen()


if __name__ == "__main__":
    main()
