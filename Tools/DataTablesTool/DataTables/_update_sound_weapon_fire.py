# -*- coding: utf-8 -*-
"""Update #Sound.xlsx (Characters+Gun as 3D) and #Weapon.xlsx (fire_sound column)."""
from __future__ import annotations

import os
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent

# spatial_blend=1 => 3D SFX
SOUND_HEADERS = [
    "id",
    "location",
    "group",
    "loop",
    "volume",
    "priority",
    "mute",
    "fade_in_seconds",
    "pitch",
    "pan_stereo",
    "spatial_blend",
    "max_distance",
    "doppler_level",
]
SOUND_TYPES = [
    "int",
    "string",
    "string",
    "bool",
    "float",
    "int",
    "bool",
    "float",
    "float",
    "float",
    "float",
    "float",
    "float",
]
SOUND_COMMENTS = [
    "id",
    "音频资源路径",
    "音频组",
    "是否循环",
    "音量",
    "优先级",
    "是否静音",
    "淡入时长(秒)",
    "音调",
    "立体声方位",
    "空间混合",
    "最大听距",
    "多普勒等级",
]

# Music 2D; Characters/Gun SFX 3D (spatial_blend=1)
SOUND_ROWS = [
    # id, location, group, loop, volume, priority, mute, fade_in, pitch, pan, spatial, max_dist, doppler
    (1001, "HomeMusic", "Music", True, 1.0, 0, False, 1.0, 1.0, 0.0, 0.0, 100.0, 1.0),
    (1002, "BGM_Battle", "Music", True, 1.0, 0, False, 1.0, 1.0, 0.0, 0.0, 100.0, 1.0),
    # Characters/Enemy
    (2001, "Enemy 01 Projectile Launch", "SFX", False, 1.0, 0, False, 0.0, 1.0, 0.0, 1.0, 100.0, 1.0),
    (2002, "Enemy 02  Projectile Launch", "SFX", False, 1.0, 0, False, 0.0, 1.0, 0.0, 1.0, 100.0, 1.0),
    (2003, "Enemy 03 Projectile Launch", "SFX", False, 1.0, 0, False, 0.0, 1.0, 0.0, 1.0, 100.0, 1.0),
    (2004, "Enemy Explosion 1", "SFX", False, 1.0, 0, False, 0.0, 1.0, 0.0, 1.0, 100.0, 1.0),
    (2005, "Enemy Explosion 2", "SFX", False, 1.0, 0, False, 0.0, 1.0, 0.0, 1.0, 100.0, 1.0),
    (2006, "Enemy Projectile Hit 1", "SFX", False, 1.0, 0, False, 0.0, 1.0, 0.0, 1.0, 100.0, 1.0),
    (2007, "Enemy Projectile Hit 2", "SFX", False, 1.0, 0, False, 0.0, 1.0, 0.0, 1.0, 100.0, 1.0),
    # Characters/Player（无 Player Projectile 文件；开火改由 Gun 表驱动）
    (2102, "Player Projectile Hit", "SFX", False, 1.0, 0, False, 0.0, 1.0, 0.0, 1.0, 100.0, 1.0),
    (2103, "Player Dead", "SFX", False, 1.0, 0, False, 0.0, 1.0, 0.0, 1.0, 100.0, 1.0),
    # Gun/
    (2201, "WeaponRifle", "SFX", False, 1.0, 0, False, 0.0, 1.0, 0.0, 1.0, 100.0, 1.0),
    (2202, "WeaponShotgun", "SFX", False, 1.0, 0, False, 0.0, 1.0, 0.0, 1.0, 100.0, 1.0),
    (2203, "WeaponRPG", "SFX", False, 1.0, 0, False, 0.0, 1.0, 0.0, 1.0, 100.0, 1.0),
    (2204, "WeaponEnergy", "SFX", False, 1.0, 0, False, 0.0, 1.0, 0.0, 1.0, 100.0, 1.0),
    (2205, "WeaponSniper", "SFX", False, 1.0, 0, False, 0.0, 1.0, 0.0, 1.0, 100.0, 1.0),
    (2206, "WeaponCrossbow", "SFX", False, 1.0, 0, False, 0.0, 1.0, 0.0, 1.0, 100.0, 1.0),
]

# weapon_key -> fire_sound id（敌人不填 = 0）
FIRE_SOUND_BY_KEY = {
    "WeaponRifle": 2201,
    "WeaponShotgun": 2202,
    "WeaponRPG": 2203,
    "WeaponEnergy": 2204,
    "WeaponSniper": 2205,
    "WeaponCrossbow": 2206,
}


def _save(wb: openpyxl.Workbook, path: Path) -> None:
    tmp = path.with_suffix(".xlsx.tmp")
    wb.save(tmp)
    try:
        os.replace(tmp, path)
        print(f"wrote {path}")
    except PermissionError:
        alt = path.with_name(path.stem + ".updated.xlsx")
        os.replace(tmp, alt)
        raise SystemExit(f"locked: {path.name}, draft -> {alt.name}")


def write_sound(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["##var", *SOUND_HEADERS])
    ws.append(["##type", *SOUND_TYPES])
    ws.append(["##group", *(["c"] * len(SOUND_HEADERS))])
    ws.append(["##", *SOUND_COMMENTS])
    for row in SOUND_ROWS:
        ws.append([None, *row])
    _save(wb, path)


def ensure_fire_sound(path: Path) -> None:
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    headers = [c.value for c in ws[1]]

    if "fire_sound" not in headers:
        if "in_shop" in headers:
            insert_at = headers.index("in_shop") + 1
        else:
            insert_at = len(headers) + 1
        ws.insert_cols(insert_at)
        ws.cell(1, insert_at).value = "fire_sound"
        ws.cell(2, insert_at).value = "int"
        ws.cell(3, insert_at).value = "c"
        ws.cell(4, insert_at).value = "开火音效(TbSound.id,0=无)"
        ws.cell(1, 1).value = "##var"
        ws.cell(2, 1).value = "##type"
        ws.cell(3, 1).value = "##group"
        ws.cell(4, 1).value = "##"
        headers = [c.value for c in ws[1]]

    key_col = headers.index("weapon_key") + 1
    fire_col = headers.index("fire_sound") + 1

    for r in range(5, ws.max_row + 1):
        key = ws.cell(r, key_col).value
        if key is None:
            continue
        ws.cell(r, fire_col).value = int(FIRE_SOUND_BY_KEY.get(str(key), 0))

    _save(wb, path)


def main() -> None:
    for folder in ("Datas", "Datas_Gen"):
        base = ROOT / folder
        if not base.is_dir():
            print(f"skip missing {folder}")
            continue
        write_sound(base / "#Sound.xlsx")
        weapon = base / "#Weapon.xlsx"
        if weapon.exists():
            ensure_fire_sound(weapon)
        else:
            print(f"skip missing {weapon}")


if __name__ == "__main__":
    main()
