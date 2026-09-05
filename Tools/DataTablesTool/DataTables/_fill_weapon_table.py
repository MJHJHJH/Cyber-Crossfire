# -*- coding: utf-8 -*-
"""Fill #Weapon.xlsx from weapon prefab dump + add WeaponId enum."""
from __future__ import annotations

import os
from pathlib import Path

import openpyxl

DATAS = Path(__file__).resolve().parent / "Datas"
WEAPON_XLSX = DATAS / "#Weapon.xlsx"
ENUMS_XLSX = DATAS / "__enums__.xlsx"

# id aligns with Contents.m_PlayerWeapons index + shop prices
WEAPONS = [
    # id, price, name, weapon_key, one_handed, auto_fire, fire_delay, recoil_speed, recoil_angle,
    # init_ammo, max_ammo, add_ammo, power_weapon_max_ammo, infinite_ammo,
    # projectile_damage, projectile_speed, projectile_range,
    # projectile_prefab, weapon_model_prefab, effect_prefab, weapon_icon, in_shop
    (0, 0, "Rifle", "WeaponRifle", False, True, 0.2, 10.0, 0.2, 80, 80, 10, 80, False, 1.0, 50.0, 30.0, "Bullet-Rifle", "WpnModel-Rifle", "WeaponFire_Pistol_1", "", True),
    (1, 350, "Shotgun", "WeaponShotgun", False, True, 0.7, 5.0, 0.2, 20, 80, 10, 80, False, 1.0, 50.0, 20.0, "Bullet-Shotgun", "WpnModel-Shotgun", "WeaponFire_Pistol_1", "", True),
    (2, 800, "RPG", "WeaponRPG", False, True, 0.8, 5.0, 0.3, 10, 20, 10, 80, False, 2.0, 20.0, 120.0, "Bullet-RPG", "WpnModel-RPG", "WeaponFire_Pistol_1", "", True),
    (3, 850, "Energy Rifle", "WeaponEnergy", False, True, 0.4, 10.0, 0.5, 20, 50, 10, 80, False, 2.0, 20.0, 30.0, "Bullet-Energy", "WpnModel-Energy", "WeaponFire_Pistol_2", "", True),
    (4, 600, "Sniper Rifle", "WeaponSniper", False, True, 0.6, 5.0, 0.5, 20, 50, 10, 80, False, 3.0, 35.0, 60.0, "Bullet-Sniper", "WpnModel-Sniper", "WeaponFire_Pistol_1", "", True),
    (5, 560, "CrossBow", "WeaponCrossbow", False, True, 0.2, 10.0, 0.2, 20, 50, 10, 80, False, 1.0, 20.0, 30.0, "Bullet-Crossbow", "WpnModel-Crossbow", "WeaponFire_Pistol_1", "", True),
    (6, 0, "Enemy Canon A", "EnemyGun_Canon_A", False, True, 5.0, 5.0, 20.0, 80, 80, 10, 80, True, 1.0, 10.0, 100.0, "enemybullet-1", "", "WeaponFire_Pistol_1", "", False),
]

HEADERS = [
    "id",
    "price",
    "name",
    "weapon_key",
    "one_handed",
    "auto_fire",
    "fire_delay",
    "recoil_speed",
    "recoil_angle",
    "init_ammo",
    "max_ammo",
    "add_ammo",
    "power_weapon_max_ammo",
    "infinite_ammo",
    "projectile_damage",
    "projectile_speed",
    "projectile_range",
    "projectile_prefab",
    "weapon_model_prefab",
    "effect_prefab",
    "weapon_icon",
    "in_shop",
]

TYPES = [
    "int",
    "int",
    "string",
    "string",
    "bool",
    "bool",
    "float",
    "float",
    "float",
    "int",
    "int",
    "int",
    "int",
    "bool",
    "float",
    "float",
    "float",
    "string",
    "string",
    "string",
    "string",
    "bool",
]

COMMENTS = [
    "id",
    "商店价格",
    "显示名",
    "业务键(拾取/兼容)",
    "单手",
    "自动开火",
    "射击间隔",
    "后坐力回弹速度",
    "后坐力角度",
    "初始弹药",
    "最大弹药",
    "补给弹药",
    "强化弹药上限",
    "无限弹药",
    "子弹伤害",
    "子弹速度",
    "子弹射程",
    "子弹Prefab location",
    "武器模型Prefab location",
    "开火特效Prefab location",
    "图标Sprite location",
    "是否出现在商店/解锁列表",
]

ENUM_ITEMS = [
    ("Rifle", "步枪", 0, "玩家步枪"),
    ("Shotgun", "霰弹枪", 1, "玩家霰弹"),
    ("RPG", "火箭筒", 2, "玩家RPG"),
    ("Energy", "能量枪", 3, "玩家能量步枪"),
    ("Sniper", "狙击枪", 4, "玩家狙击"),
    ("Crossbow", "弩", 5, "玩家弩"),
    ("EnemyCanonA", "敌方加农", 6, "EnemyWeapons/EnemyGun_Canon_A"),
]


def _remove_excel_lock(path: Path) -> None:
    lock = path.parent / f"~${path.name}"
    if lock.exists():
        try:
            lock.unlink()
            print(f"removed lock {lock.name}")
        except OSError as ex:
            print(f"warn: cannot remove lock {lock}: {ex}")


def write_weapon_xlsx() -> None:
    _remove_excel_lock(WEAPON_XLSX)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # row1 ##var
    ws.append(["##var", *HEADERS])
    # row2 ##type
    ws.append(["##type", *TYPES])
    # row3 ##group
    ws.append(["##group", *(["c"] * len(HEADERS))])
    # row4 ## comments
    ws.append(["##", *COMMENTS])
    for row in WEAPONS:
        ws.append([None, *row])

    # write to temp then replace (avoids Excel share lock on open handle)
    tmp = WEAPON_XLSX.with_suffix(".xlsx.tmp")
    wb.save(tmp)
    try:
        os.replace(tmp, WEAPON_XLSX)
    except PermissionError:
        alt = DATAS / "#Weapon.filled.xlsx"
        os.replace(tmp, alt)
        raise SystemExit(
            f"ERROR: #Weapon.xlsx is locked (close Excel). Wrote draft to {alt.name}"
        )
    print(f"wrote {WEAPON_XLSX} rows={len(WEAPONS)}")


def write_enums_xlsx() -> None:
    _remove_excel_lock(ENUMS_XLSX)
    wb = openpyxl.load_workbook(ENUMS_XLSX)
    ws = wb.active

    # Clear old data rows (keep header 1-3)
    for r in range(ws.max_row, 3, -1):
        ws.delete_rows(r)

    # Enum header row + first item
    # columns: A=##/empty, B=full_name, C=flags, D=unique, E=group, F=comment, G=tags,
    # H=name, I=alias, J=value, K=comment, L=tags
    first = ENUM_ITEMS[0]
    ws.append(
        [
            None,
            "WeaponId",
            False,
            True,
            None,
            "武器Id(与TbWeapon.id一致)",
            None,
            first[0],
            first[1],
            first[2],
            first[3],
            None,
        ]
    )
    for item in ENUM_ITEMS[1:]:
        ws.append(
            [
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                item[0],
                item[1],
                item[2],
                item[3],
                None,
            ]
        )

    tmp = ENUMS_XLSX.with_suffix(".xlsx.tmp")
    wb.save(tmp)
    try:
        os.replace(tmp, ENUMS_XLSX)
    except PermissionError:
        alt = DATAS / "__enums__.filled.xlsx"
        os.replace(tmp, alt)
        raise SystemExit(
            f"ERROR: __enums__.xlsx is locked (close Excel). Wrote draft to {alt.name}"
        )
    print(f"wrote WeaponId enum items={len(ENUM_ITEMS)}")


def main() -> None:
    write_weapon_xlsx()
    write_enums_xlsx()


if __name__ == "__main__":
    main()
